#IMPORTAMOS LAS LIBRERIAS DE INTERÉS
#To install openpyxl type 'pip install openpyxl' or 'pip3 install openpyxl' in your terminal
from openpyxl import load_workbook
from tkinter import *
import random

'''IMPORTAMOS LOS DATOS DE NUESTRAS HOJAS DE EXCEL'''
#Cargamos el archivo de datos
data_file = load_workbook("datos.xlsx")
#Obtenemos los nombres de las hojas
hojas=data_file.sheetnames
#Abrimos cada hoja en una variable diferente
key=data_file[hojas[0]]
data=data_file[hojas[1]]
#Crear la matriz de preguntas y respuestas en base al archivo de excel
pre_res=[]
for i, row in enumerate(key):
    if i==0:
        continue
    pre_res.append([row[0].value,row[1].value])

#Creamos la matriza de datos de los usuarios
users=[[],[],[],[]]
for i,  row in enumerate(data):
    if i==0:
        continue
    users[0].append(row[0].value)
    users[1].append(row[1].value)
    users[2].append(row[2].value)
    users[3].append(row[3].value)

#Creamos una matriz para nuevos usuarios
new_users=[]

'''FUNCIÓN PARA QUE SELECCIONEMOS LAS PREGUNTAS QUE QUEREMOS QUE NUESTRO USUARIO RESPONDA'''
def seleccion_preguntas():
    #Determinamos la cantidad de preguntas que hay
    cant_preguntas=len(pre_res)

    #Creamos una lista con los numeros del 0 hasta la cantidad de preguntas disponibles
    seleccion=[]
    for i in range(0, cant_preguntas):
        seleccion.append(i)

    #Quitamos tantos valores cómo sean necesarios para solo tener 10
    for i in range(0, cant_preguntas-10):
        seleccion.remove(random.choice(seleccion))
    
    #Creamos listas para guardar las preguntas y respuestas seleccionadas
    preguntas=[]
    respuestas=[]
    #Guardamos las preguntas y respuestas correspondientes
    for i in seleccion:
        preguntas.append(pre_res[i][0])
        respuestas.append(pre_res[i][1])
    
    return preguntas, respuestas

def create_rows(window, rows):
    for i in range(0, rows):
        window.rowconfigure(i, weight=1)

#data_file.save("datos.xlsx")

'''TKINTER'''
#Ventana madre
main=Tk()

#Imagenes
logo=PhotoImage(file="logo.png")
bg=PhotoImage(file="fondo1.png")

#Ventana de inicio
main.title("Quiz PLI")
main.geometry("500x800")
main.config(bg="#373D64")
main.columnconfigure(0, weight=1)
main.rowconfigure(0, weight=1)
main.rowconfigure(1, weight=1)
main.rowconfigure(2, weight=1)
main.resizable(0, 0)

#Ventana de Ingreso
w1=Toplevel(main)
w1.title("Ingreso")
w1.geometry("500x800")
w1.config(bg="#373D64")
w1.resizable(0, 0)
w1.columnconfigure(0, weight=1)
create_rows(w1, 4)
w1.iconify()

#Ventana de Registro
w2=Toplevel(main)
w2.title("Registro")
w2.geometry("500x800")
w2.config(bg="#373D64")
w2.resizable(0, 0)
w2.columnconfigure(0, weight=1)
create_rows(w2, 5)
w2.iconify()

#Ventana de Bienvenida y Datos
w3=Toplevel(main)
w3.title("Quiz")
w3.geometry("500x800")
w3.config(bg="#373D64")
w3.resizable(0, 0)
w3.columnconfigure(0, weight=1)
create_rows(w3, 4)
w3.iconify()

#Ventana de Preguntas
w4=Toplevel(main)
w4.title("Preguntas")
w4.geometry("500x800")
w4.config(bg="#373D64")
w4.resizable(0, 0)
w4.columnconfigure(0, weight=1)
create_rows(w4, 4)
w4.iconify()

#Ventana de Resultados
w5=Toplevel(main)
w5.title("Resultados")
w5.geometry("500x800")
w5.config(bg="#373D64")
w5.resizable(0, 0)
w5.columnconfigure(0, weight=1)
create_rows(w5, 4)
w5.iconify()

'''Formatos'''
#Labels
label_f = {"bg": ("#373D64"), "fg": ("White")}
#Botones
button_f = {"fg": ("#373D64")}
#Texto
title = {"font": (None, 40)}
subtitle = {"font": (None, 30)}
text = {"font": (None, 20)}

'''Funciones'''
def abrir_ingreso():
    main.iconify()
    w1.deiconify()

def abrir_registro():
    main.iconify()
    w2.deiconify()

def ingresar():
    global matricula, contrasena, nombre, score
    #Obtener los datos de la ventana
    matricula=e1_w1.get()
    contrasena=e2_w1.get()
    if matricula in users[0]:
        ind=users[0].index(matricula)
        if contrasena==users[1][ind]:
            e1_w1.delete(0, END)
            e2_w1.delete(0, END)
            nombre=users[2][ind]
            score=users[3][ind]
            l1_w3.configure(text=f"Bienvenido, {nombre}")
            if score=="-":
                l2_w3.configure(text=f"Todavía no has\nhecho el quiz")
            else:
                l2_w3.configure(text=f"Calificación\nactual: {score}/100")
            w1.iconify()
            w3.deiconify()
        else:
            l4_w1.configure(text="Contraseña incorrecta")
    else:
        l4_w1.configure(text="Esta matrícula no está en la base de datos")

def registrar():
    global matricula, contrasena, nombre, score, new_users
    matricula=e1_w2.get()
    contrasena=e2_w2.get()
    nombre=e3_w2.get()
    score="-"
    if matricula[:5]=="A0170" and len(matricula) == 9 and len(contrasena)>3 and nombre!="":
        if matricula not in users[0]:
            e1_w2.delete(0, END)
            e2_w2.delete(0, END)
            e3_w2.delete(0, END)
            users[0].append(matricula)
            users[1].append(contrasena)
            users[2].append(nombre)
            users[3].append(score)
            new_users.append([matricula, contrasena, nombre, score])
            l1_w3.configure(text=f"Bienvenido, {nombre}")
            l2_w3.configure(text="Todavía no has\nhecho el quiz")
            w2.iconify()
            w3.deiconify()
        else:
            l5_w2.configure(text="Esta matrícula ya está en la base de datos")
    elif matricula[:5]!="A0170" or len(matricula)!=9:
        l5_w2.configure(text="El formate de matrícula\nes: 'A0170XXXX'")
    elif len(contrasena)<=3:
        l5_w2.configure(text="La longitud de la contraseña\nno puede ser mayor a 3 caracteres")
    else:
        l5_w2.configure(text="El nombre no puede estar vacío")

def empezar_examen():
    global preguntas, respuestas, indice_pregunta, correct
    preguntas, respuestas=seleccion_preguntas()
    correct=0
    indice_pregunta=0
    l1_w4.configure(text=f"{indice_pregunta+1}.- {preguntas[indice_pregunta]}")
    w3.iconify()
    w4.deiconify()

def cerrar_sesion():
    global matricula, contrasena, nombre, score
    matricula=""
    contrasena=""
    nombre=""
    score=""
    l1_w3.configure(text=f"Bienvenido")
    l2_w3.configure(text="Todavía no has\nhecho el quiz")
    w3.iconify()
    main.deiconify()

def avanzar_preguntas():
    global preguntas, respuestas, indice_pregunta, correct
    prompt=e1_w4.get()
    e1_w4.delete(0, END)
    if prompt == respuestas[indice_pregunta]:
        correct+=1
    indice_pregunta+=1
    l1_w4.configure(text=f"{indice_pregunta+1}.- {preguntas[indice_pregunta]}")
    if indice_pregunta == len(respuestas)-1:
        b1_w4.configure(text="Terminar",command=terminar_preguntas)
    print()


def terminar_preguntas():
    global matricula, score, respuestas, correct
    prompt=e1_w4.get()
    e1_w4.delete(0, END)
    if prompt == respuestas[indice_pregunta]:
        correct+=1
    if score == "-":
        score=0
    resultado=correct*10
    if resultado > 70:
        l1_w5.configure(text="Has aprobado")
    else:
        l1_w5.configure(text="Has reprobado")
    l2_w5.configure(text=f"Tú calificación es\nde: {resultado}/100")
    if int(score) < resultado:
        score=str(resultado)
    index_user=users[0].index(matricula)
    users[3][index_user]=score
    data[index_user+2][3].value=score
    l3_w5.configure(text=f"Tú calificación es\nd: {score}/100")
    w4.iconify()
    w5.deiconify()

def repetir():
    global preguntas, respuestas, indice_pregunta, correct
    preguntas, respuestas=seleccion_preguntas()
    correct=0
    indice_pregunta=0
    l1_w4.configure(text=f"{indice_pregunta+1}.- {preguntas[indice_pregunta]}")
    b1_w4.configure(text="Siguiente",command=avanzar_preguntas)
    w5.iconify()
    w4.deiconify()

def regresar():
    l2_w3.configure(text=f"Calificación\nactual: {score}/100")
    w5.iconify()
    w3.deiconify()

def salir():
    if new_users!=[]:
        for i in new_users:
            data.append(i)
    data_file.save("datos.xlsx")
    main.destroy()

#Colocar fondo
bgl_main=Label(main, image=bg, **label_f).place(x=-1, y=-1)
bgl_w1=Label(w1, image=bg, **label_f).place(x=-1, y=-1)
bgl_w2=Label(w2, image=bg, **label_f).place(x=-1, y=-1)
bgl_w3=Label(w3, image=bg, **label_f).place(x=-1, y=-1)
bgl_w4=Label(w4, image=bg, **label_f).place(x=-1, y=-1)
bgl_w5=Label(w5, image=bg, **label_f).place(x=-1, y=-1)

#Elementos Ventana Inicio
l1_main=Label(main, image=logo, **label_f).pack(pady=40)
b1_main=Button(main, text="Ingresar", **subtitle, **button_f, command=abrir_ingreso).pack(pady=40)
b2_main=Button(main, text="Registrarse", **subtitle, **button_f, command=abrir_registro).pack(pady=40)
b3_main=Button(main, text="Salir", **subtitle, **button_f, command=salir).pack(pady=40)

#Elementos Ventana de Ingreso
l1_w1=Label(w1, text="Ingresar", **title, **label_f)
l1_w1.grid(row=0, column=0)
l2_w1=Label(w1, text="Matrícula:", **subtitle, **label_f)
l2_w1.grid(row=1, column=0, sticky=N)
l3_w1=Label(w1, text="Contraseña:", **subtitle, **label_f)
l3_w1.grid(row=2, column=0, sticky=N)
l4_w1=Label(w1, text="", **text, **label_f)
l4_w1.grid(row=2, column=0, sticky=N, pady=100)
e1_w1=Entry(w1, **subtitle)
e1_w1.grid(row=1, column=0, sticky=N, pady=50)
e2_w1=Entry(w1, **subtitle)
e2_w1.grid(row=2, column=0, sticky=N, pady=50)
b1_w1=Button(w1, text="Ingresar", **subtitle, **button_f, command=ingresar)
b1_w1.grid(row=3, column=0)

#Elementos Ventana de Registro
l1_w2=Label(w2, text="Registro", **title, **label_f)
l1_w2.grid(row=0, column=0)
l2_w2=Label(w2, text="Matrícula:", **subtitle, **label_f)
l2_w2.grid(row=1, column=0, sticky=N)
l3_w2=Label(w2, text="Contraseña:", **subtitle, **label_f)
l3_w2.grid(row=2, column=0, sticky=N)
l4_w2=Label(w2, text="Nombre:", **subtitle, **label_f)
l4_w2.grid(row=3, column=0, sticky=N)
l5_w2=Label(w2, text="", **text, **label_f)
l5_w2.grid(row=3, column=0, sticky=N, pady=100)
e1_w2=Entry(w2, **subtitle)
e1_w2.grid(row=1, column=0, sticky=N, pady=50)
e2_w2=Entry(w2, **subtitle)
e2_w2.grid(row=2, column=0, sticky=N, pady=50)
e3_w2=Entry(w2, **subtitle)
e3_w2.grid(row=3, column=0, sticky=N, pady=50)
b1_w2=Button(w2, text="Registrar", **subtitle, **button_f, command=registrar)
b1_w2.grid(row=4, column=0)

#Elementos Ventana de Bienvenida y Datos
l1_w3=Label(w3, text="Bienvenido, usuario", **title, **label_f)
l1_w3.grid(row=0, column=0)
l2_w3=Label(w3, text="Todavía no has\nhecho el quiz", **subtitle, **label_f)
l2_w3.grid(row=1, column=0)
b1_w3=Button(w3, text="Realizar examen", **subtitle, **button_f, command=empezar_examen)
b1_w3.grid(row=2, column=0)
b2_w3=Button(w3, text="Cerrar sesión", **subtitle, **button_f, command=cerrar_sesion)
b2_w3.grid(row=3, column=0)

#Elementos Ventana Preguntas
l1_w4=Label(w4, text="1.-¿Pregunta?", **text, **label_f)
l1_w4.grid(row=0, column=0)
l2_w4=Label(w4, text="Respuesta:", **subtitle, **label_f)
l2_w4.grid(row=1, column=0)
e1_w4=Entry(w4, **subtitle)
e1_w4.grid(row=1, column=0, sticky=S, pady=50)
b1_w4=Button(w4, text="Siguiente", **subtitle, **button_f, command=avanzar_preguntas)
b1_w4.grid(row=2, column=0)

#Elementos Ventana de Bienvenida y Datos
l1_w5=Label(w5, text="Resultado", **title, **label_f)
l1_w5.grid(row=0, column=0)
l2_w5=Label(w5, text="Tú calificación\nes de: -/100", **subtitle, **label_f)
l2_w5.grid(row=1, column=0, sticky=N, pady=50)
l3_w5=Label(w5, text="Tú calificación más\nalta es de: -/100", **subtitle, **label_f)
l3_w5.grid(row=1, column=0, sticky=S, pady=50)
b1_w5=Button(w5, text="Repetir examen", **subtitle, **button_f, command=repetir)
b1_w5.grid(row=2, column=0)
b2_w5=Button(w5, text="Regresar", **subtitle, **button_f, command=regresar)
b2_w5.grid(row=3, column=0)

main.mainloop()